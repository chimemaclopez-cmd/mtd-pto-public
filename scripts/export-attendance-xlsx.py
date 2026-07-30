#!/usr/bin/env python3
# Regenerates each rep's personal attendance record whenever attendance.json changes.
# Called automatically by zendesk-proxy.js after every attendance save (manual entry,
# PTO approval/reversal, team-lead sync) - one-way mirror (portal -> Excel), same
# convention as export-roster-xlsx.py.
#
# Layout: Lofty TSR/Reps/<Employee Name>/Attendance/<Year>/Attendance.xlsx - one
# workbook per rep per year (bounds tab growth), one sheet per month that actually has
# an entry (no blank future-month tabs, since this file is never edited directly - the
# portal is authoritative and this is read-only). Never deletes a rep's folder if they
# leave the roster later - only creates/updates.
import json
import os
import re
import sys
from datetime import date as date_cls
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROSTER_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'roster.json')
ATTENDANCE_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'attendance.json')
REPS_ROOT = '/Users/mac/Library/CloudStorage/OneDrive-MoatableInc/Lofty PH Repository - Documents/Lofty TSR/Reps'

STATUS_LABELS = {
    'ONSITE': 'Onsite', 'WFH': 'Work From Home', 'LATE': 'Late', 'RD': 'Rest Day',
    'PTO': 'PTO', 'PARTIAL_PTO': 'Partial PTO', 'SL': 'Sick Leave', 'SL-HD': 'Sick Leave (Half Day)',
    'EL': 'Emergency Leave', 'EL-HD': 'Emergency Leave (Half Day)', 'NCNS': 'No Call No Show',
    'A': 'Absent', 'BL': 'Bereavement Leave', 'SUSPENDED': 'Suspended',
}
WEEKDAYS = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def sanitize_folder_name(name, fallback):
    cleaned = re.sub(r'[/\\:*?"<>|]', '-', str(name or '')).strip()
    return cleaned or fallback


def weekday_of(date_str):
    y, m, d = (int(x) for x in date_str.split('-'))
    return WEEKDAYS[date_cls(y, m, d).isoweekday() % 7]


def resolve_code_for_date(attendance, email, date_str):
    auto = (attendance.get('autoEntries') or {}).get(email, {}).get(date_str)
    if auto:
        status = auto.get('status') if isinstance(auto, dict) else auto
        return status, auto if isinstance(auto, dict) else None, 'Approved PTO'
    month = date_str[:7]
    matches = sorted(
        [(k, v) for k, v in (attendance.get('periods') or {}).items() if k.startswith(month + '|') and k.split('|')[1] >= date_str],
        key=lambda kv: kv[0], reverse=True,
    )
    for _, period in matches:
        value = (period.get(email) or {}).get(date_str)
        if value:
            status = value.get('status') if isinstance(value, dict) else value
            return status, value if isinstance(value, dict) else None, 'Manual Entry'
    return None, None, None


def collect_dates_for_employee(attendance, email):
    dates = set()
    for period in (attendance.get('periods') or {}).values():
        for d, v in (period.get(email) or {}).items():
            if v:
                dates.add(d)
    for d, v in ((attendance.get('autoEntries') or {}).get(email) or {}).items():
        if v:
            dates.add(d)
    return sorted(dates)


def build_month_sheet(ws, email, attendance, dates_in_month):
    body_font = Font(name='Calibri', size=11)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2E46B8')
    thin = Side(style='thin', color='DFE2EA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Date', 'Weekday', 'Status', 'Status Label', 'Detail', 'Source']
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.freeze_panes = 'A2'

    row_idx = 2
    for date_str in dates_in_month:
        status, extra, source = resolve_code_for_date(attendance, email, date_str)
        if not status:
            continue
        detail = ''
        if isinstance(extra, dict):
            if extra.get('minutesLate'):
                detail = f"{extra['minutesLate']} min late"
            elif extra.get('requestId'):
                detail = extra['requestId'] + (f" ({extra['durationMinutes']} min)" if extra.get('durationMinutes') else '')
        row = [date_str, weekday_of(date_str), status, STATUS_LABELS.get(status, status), detail, source or '']
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical='center')
        row_idx += 1

    widths = [14, 14, 12, 22, 26, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if row_idx > 2:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row_idx - 1}'


def save_with_retry(wb, out_path, attempts=3):
    # OneDrive's sync/File Provider layer occasionally stalls a write with a transient
    # timeout under repeated rapid file operations - retry rather than let one flaky
    # write fail the whole export.
    import time
    last_error = None
    for attempt in range(attempts):
        try:
            wb.save(out_path)
            return
        except (TimeoutError, OSError) as e:
            last_error = e
            if attempt < attempts - 1:
                time.sleep(1.5 * (attempt + 1))
    raise last_error


def main():
    # Optional CLI args: specific employee emails to regenerate. Passing only the
    # rep(s) actually affected by a save (the common case) avoids rewriting all 42
    # reps' files - and their folders' worth of OneDrive I/O - on every single change.
    # With no args, regenerates everyone (used for the first run / a full resync).
    target_emails = {e.strip().lower() for e in sys.argv[1:] if e.strip()}

    with open(ROSTER_PATH) as f:
        roster = json.load(f)
    with open(ATTENDANCE_PATH) as f:
        attendance = json.load(f)

    # First pass over the FULL roster: compute every employee's folder name so
    # collision disambiguation is correct even when only regenerating a subset.
    used_names = {}
    folder_by_email = {}
    for record in roster['records']:
        email = record.get('employeeEmail')
        if not email:
            continue
        name = record.get('employeeName') or email
        folder_name = sanitize_folder_name(name, email)
        if folder_name in used_names and used_names[folder_name] != email:
            folder_name = f"{folder_name} ({email})"
        used_names[folder_name] = email
        folder_by_email[email] = folder_name

    written = 0
    for record in roster['records']:
        email = record.get('employeeEmail')
        if not email:
            continue
        if target_emails and email.lower() not in target_emails:
            continue
        folder_name = folder_by_email[email]

        attendance_root = os.path.join(REPS_ROOT, folder_name, 'Attendance')
        dates = collect_dates_for_employee(attendance, email)
        by_year = {}
        for d in dates:
            by_year.setdefault(d[:4], []).append(d)

        if not by_year:
            # No history at all yet - still give the rep a current-year file to start from.
            by_year[str(date_cls.today().year)] = []

        for year, dates_in_year in by_year.items():
            year_dir = os.path.join(attendance_root, year)
            os.makedirs(year_dir, exist_ok=True)
            wb = Workbook()
            wb.remove(wb.active)
            by_month = {}
            for d in dates_in_year:
                by_month.setdefault(d[5:7], []).append(d)
            for month_num in sorted(by_month.keys()):
                sheet_name = f"{MONTH_NAMES[int(month_num) - 1]} {year}"
                ws = wb.create_sheet(sheet_name)
                build_month_sheet(ws, email, attendance, sorted(by_month[month_num]))
            if not wb.sheetnames:
                ws = wb.create_sheet('Attendance')
                build_month_sheet(ws, email, attendance, [])
            save_with_retry(wb, os.path.join(year_dir, 'Attendance.xlsx'))
        written += 1

    print(f'Exported attendance records for {written} rep(s) under {REPS_ROOT}')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'attendance export failed: {e}', file=sys.stderr)
        sys.exit(1)
