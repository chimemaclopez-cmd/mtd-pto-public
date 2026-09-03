#!/usr/bin/env python3
# Regenerates a company-wide, consolidated attendance workbook per year whenever
# attendance.json changes - called automatically by zendesk-proxy.js alongside the
# per-rep export (export-attendance-xlsx.py). Unlike that one (one workbook per rep,
# chronological rows), this is one workbook per YEAR covering every rep at once, laid
# out as a matrix (rows=reps, columns=days) so a reviewer can scan the whole team for a
# period at a glance, plus a Summary sheet with computed attendance-reliability rates
# and flagged at-risk employees - both recomputed fresh from source data every run, so
# they're always current with no manual upkeep.
#
# Layout: Lofty TSR/Attendance/<Year>.xlsx - a "Summary" sheet first, then one sheet
# per calendar month that has any attendance data across any rep.
#
# The attendance-rate formula (present / eligible * 100) mirrors
# server/pto-logic.js's computeAttendanceForRange() exactly - the same formula already
# shown as "Attendance %" in KPI Scores and Evaluation PDFs - so this number means the
# same thing everywhere in the portal. The schedule-resolution helpers below
# (roster_active_on/schedule_for_date/attendance_code_on_date) are a direct port of that
# same file's rosterActiveOn/scheduleForDate/attendanceCodeOnDate - duplicated here for
# the same Python/Node boundary reason as ATTENDANCE_CODES elsewhere in this repo.
import calendar
import json
import os
import re
import sys
import time
from datetime import date as date_cls, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROSTER_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'roster.json')
ATTENDANCE_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'attendance.json')
SCHEDULES_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'schedules.json')
OUT_ROOT = '/Users/mac/Library/CloudStorage/OneDrive-MoatableInc/Lofty PH Repository - Documents/Lofty TSR/Attendance'

WEEKDAYS = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
STATUS_COLORS = {
    'ONSITE': 'FDF3D9', 'LATE': 'FDE3CC', 'WFH': 'DCF3E6', 'RD': 'EEF0F5',
    'PTO': 'EDE9FE', 'PARTIAL_PTO': 'EDE9FE',
    'SL': 'FBE4E4', 'EL': 'FBE4E4', 'SL-HD': 'FBE4E4', 'EL-HD': 'FBE4E4', 'BL': 'FBE4E4',
    'NCNS': 'F6C9C9', 'A': 'F6C9C9', 'SUSPENDED': 'F6C9C9',
}
NAVY = '2E46B8'
GREEN = 'E3F7EC'
AMBER = 'FFF2CC'
RED = 'FCE8E8'
TODAY = date_cls.today().isoformat()
THIN = Side(style='thin', color='DFE2EA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean_email(value):
    return str(value or '').strip().lower()


def valid_date(value):
    return bool(re.match(r'^\d{4}-\d{2}-\d{2}$', str(value or '')))


def plus_days(date_str, days):
    y, m, d = (int(x) for x in date_str.split('-'))
    return (date_cls(y, m, d) + timedelta(days=days)).isoformat()


def date_range(start, end):
    if not valid_date(start) or not valid_date(end) or start > end:
        return []
    out, d = [], start
    while d <= end:
        out.append(d)
        d = plus_days(d, 1)
    return out


def weekday_of(date_str):
    y, m, d = (int(x) for x in date_str.split('-'))
    return WEEKDAYS[date_cls(y, m, d).isoweekday() % 7]


def roster_active_on(employee, date_str):
    if not employee or employee.get('active') is False:
        return False
    if employee.get('employmentStatus') in ('Resigned', 'Terminated', 'Inactive'):
        return False
    hire, sep = employee.get('hireDate'), employee.get('separationDate')
    if hire and hire > date_str:
        return False
    if sep and sep < date_str:
        return False
    return True


def schedule_for_date(schedule_data, email, date_str):
    email = clean_email(email)
    all_records = [x for x in (schedule_data.get('schedules') or [])
                    if clean_email(x.get('employeeEmail')) == email and (x.get('active') or x.get('effectiveTo'))]
    records = sorted(
        [x for x in all_records if x.get('effectiveFrom', '') <= date_str and (not x.get('effectiveTo') or x.get('effectiveTo', '') >= date_str)],
        key=lambda x: x.get('effectiveFrom', ''), reverse=True,
    )
    fallback = sorted(all_records, key=lambda x: x.get('effectiveFrom', ''), reverse=True)[0] if (not records and all_records) else None
    record = records[0] if records else fallback
    weekday = weekday_of(date_str)
    base = (record or {}).get('weekly', {}).get(weekday) if record else None
    overrides = sorted(
        [x for x in (schedule_data.get('overrides') or []) if clean_email(x.get('employeeEmail')) == email and x.get('date') == date_str],
        key=lambda x: str(x.get('updatedAt') or ''), reverse=True,
    )
    override = overrides[0] if overrides else None
    template = {**(base or {}), **override, 'assignments': override.get('assignments') or (base or {}).get('assignments') or {}} if override else base
    return {'template': template, 'missingSchedule': not base and not override}


def attendance_code_on_date(attendance, email, date_str):
    auto = (attendance.get('autoEntries') or {}).get(email, {}).get(date_str)
    if auto:
        return auto.get('status') if isinstance(auto, dict) else auto
    month = date_str[:7]
    matches = sorted(
        [(k, v) for k, v in (attendance.get('periods') or {}).items() if k.startswith(month + '|') and k.split('|')[1] >= date_str],
        key=lambda kv: kv[0], reverse=True,
    )
    for _, period in matches:
        value = (period.get(email) or {}).get(date_str)
        if value:
            return value.get('status') if isinstance(value, dict) else value
    return ''


def compute_attendance_rate(roster_by_email, schedules, attendance, email, start_date, end_date):
    employee = roster_by_email.get(clean_email(email))
    if not employee:
        return None
    present = eligible = ncns = 0
    absence = 0.0
    for d in date_range(start_date, end_date):
        if not roster_active_on(employee, d):
            continue
        resolved = schedule_for_date(schedules, email, d)
        if resolved['missingSchedule'] or (resolved['template'] or {}).get('off'):
            continue
        code = attendance_code_on_date(attendance, email, d)
        if code == 'PTO':
            continue
        eligible += 1
        if code == 'PARTIAL_PTO' or not code:
            continue
        if code in ('ONSITE', 'WFH', 'LATE'):
            present += 1
        else:
            if code == 'NCNS':
                ncns += 1
            absence += 0.5 if code in ('SL-HD', 'EL-HD') else 1
    return {'present': present, 'eligible': eligible, 'ncns': ncns, 'absence': absence,
            'rate': (present / eligible * 100) if eligible else None}


def rate_fill(rate):
    if rate is None:
        return None
    return PatternFill('solid', fgColor=GREEN if rate >= 95 else AMBER if rate >= 90 else RED)


def build_month_sheet(ws, year, month_num, roster_rows, schedules, attendance, roster_by_email):
    days_in_month = calendar.monthrange(int(year), month_num)[1]
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor=NAVY)

    headers = ['Employee', 'Team Lead'] + [str(d) for d in range(1, days_in_month + 1)] + ['Month %']
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    month_start = f'{year}-{month_num:02d}-01'
    # Capped at today for the CURRENT/a future month, so an in-progress month's remaining
    # (not-yet-happened) days don't count as unexplained absences in the denominator - the
    # matrix columns below still show every day of the month regardless, just blank past today.
    month_end = min(f'{year}-{month_num:02d}-{days_in_month:02d}', TODAY)

    row_idx = 2
    for emp in roster_rows:
        email = emp.get('employeeEmail')
        ws.cell(row=row_idx, column=1, value=emp.get('employeeName') or email).border = BORDER
        ws.cell(row=row_idx, column=2, value=emp.get('teamLeadName') or '').border = BORDER
        for day in range(1, days_in_month + 1):
            date_str = f'{year}-{month_num:02d}-{day:02d}'
            code = attendance_code_on_date(attendance, email, date_str)
            cell = ws.cell(row=row_idx, column=2 + day)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
            if code:
                cell.value = code
                cell.font = Font(size=8)
                fill_color = STATUS_COLORS.get(code)
                if fill_color:
                    cell.fill = PatternFill('solid', fgColor=fill_color)
        stats = compute_attendance_rate(roster_by_email, schedules, attendance, email, month_start, month_end)
        rate = stats['rate'] if stats else None
        rate_cell = ws.cell(row=row_idx, column=len(headers), value=(round(rate, 1) if rate is not None else None))
        rate_cell.border = BORDER
        rate_cell.alignment = Alignment(horizontal='center')
        rate_cell.font = Font(bold=True)
        fill = rate_fill(rate)
        if fill:
            rate_cell.fill = fill
        row_idx += 1

    ws.freeze_panes = 'C2'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    for day in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(2 + day)].width = 5
    ws.column_dimensions[get_column_letter(len(headers))].width = 10
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row_idx - 1}'


def build_summary_sheet(ws, year, roster_rows, schedules, attendance, roster_by_email, months_with_data):
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor=NAVY)
    bold = Font(bold=True)

    ws.cell(row=1, column=1, value=f'{year} Attendance Reliability Summary').font = Font(size=16, bold=True)
    ws.cell(row=2, column=1, value=(
        'Attendance Rate = Present Days ÷ Eligible Scheduled Days (approved PTO, rest days, and days '
        'outside the employee’s active employment window are excluded from both sides). This is the same '
        'formula used for the Attendance component in KPI Scores and Evaluation PDFs, so the number means the '
        'same thing everywhere in the portal. Recomputed fresh from live data every time this file regenerates.'
    )).font = Font(size=9, italic=True, color='5B6274')
    ws.merge_cells('A2:Q2')
    ws.row_dimensions[2].height = 28
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    rows = []
    for emp in roster_rows:
        email = emp.get('employeeEmail')
        # "Year Rate" aggregates present/eligible across only the tracked months (months_with_data),
        # not a blind Jan-Dec span - otherwise months before the portal was in use for this team
        # would resolve a valid schedule with no recorded attendance code, silently counting every
        # one of those days as an unexplained absence and crushing the rate (confirmed live: this
        # produced an 18% company-wide rate before the fix, for a team that only just started using
        # the portal in July).
        year_present = year_eligible = year_ncns = 0
        year_absence = 0.0
        month_rates = [None] * 12
        for m in range(1, 13):
            dim = calendar.monthrange(int(year), m)[1]
            m_end = min(f'{year}-{m:02d}-{dim:02d}', TODAY)  # see build_month_sheet's comment on month_end
            ms = compute_attendance_rate(roster_by_email, schedules, attendance, email, f'{year}-{m:02d}-01', m_end)
            if m in months_with_data and ms:
                month_rates[m - 1] = ms['rate']
                year_present += ms['present']
                year_eligible += ms['eligible']
                year_ncns += ms['ncns']
                year_absence += ms['absence']
        rows.append({
            'name': emp.get('employeeName') or email, 'teamLead': emp.get('teamLeadName') or '',
            'rate': (year_present / year_eligible * 100) if year_eligible else None,
            'ncns': year_ncns, 'absence': year_absence, 'eligible': year_eligible,
            'month_rates': month_rates,
        })

    tracked = [r for r in rows if r['eligible'] > 0]
    rates = [r['rate'] for r in tracked if r['rate'] is not None]
    company_avg = sum(rates) / len(rates) if rates else None
    below_90 = [r for r in tracked if r['rate'] is not None and r['rate'] < 90]
    total_ncns = sum(r['ncns'] for r in rows)

    stat_row = 4
    for label, value in [
        ('Employees Tracked This Year', len(tracked)),
        ('Company-Wide Average Attendance Rate', f"{company_avg:.1f}%" if company_avg is not None else '—'),
        ('Employees Below 90% (Needs Attention)', len(below_90)),
        ('Total NCNS Incidents This Year', total_ncns),
    ]:
        ws.cell(row=stat_row, column=1, value=label).font = bold
        ws.cell(row=stat_row, column=3, value=value)
        stat_row += 1

    if below_90:
        stat_row += 1
        ws.cell(row=stat_row, column=1, value='⚠ Needs Attention (below 90% for the year, lowest first):').font = Font(bold=True, color='C2313A')
        stat_row += 1
        for r in sorted(below_90, key=lambda x: x['rate']):
            ws.cell(row=stat_row, column=1, value=f"{r['name']} ({r['teamLead']}) — {r['rate']:.1f}%, {r['ncns']} NCNS")
            stat_row += 1

    table_start = stat_row + 2
    headers = ['Employee', 'Team Lead', 'Year Rate'] + MONTH_NAMES + ['NCNS', 'Absence Days']
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    row_idx = table_start + 1
    for r in sorted(rows, key=lambda x: (x['rate'] is None, x['rate'] if x['rate'] is not None else 0)):
        ws.cell(row=row_idx, column=1, value=r['name']).border = BORDER
        ws.cell(row=row_idx, column=2, value=r['teamLead']).border = BORDER
        rate_cell = ws.cell(row=row_idx, column=3, value=(round(r['rate'], 1) if r['rate'] is not None else None))
        rate_cell.font = bold
        rate_cell.border = BORDER
        rate_cell.alignment = Alignment(horizontal='center')
        fill = rate_fill(r['rate'])
        if fill:
            rate_cell.fill = fill
        for i, mr in enumerate(r['month_rates']):
            c = ws.cell(row=row_idx, column=4 + i, value=(round(mr, 1) if mr is not None else None))
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=9)
            c.border = BORDER
        ws.cell(row=row_idx, column=16, value=r['ncns']).border = BORDER
        ws.cell(row=row_idx, column=16).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=17, value=round(r['absence'], 1)).border = BORDER
        ws.cell(row=row_idx, column=17).alignment = Alignment(horizontal='center')
        row_idx += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    for col in range(3, 18):
        ws.column_dimensions[get_column_letter(col)].width = 8
    ws.freeze_panes = f'C{table_start + 1}'
    ws.auto_filter.ref = f'A{table_start}:{get_column_letter(len(headers))}{row_idx - 1}'


def save_with_retry(wb, out_path, attempts=3):
    for attempt in range(attempts):
        try:
            wb.save(out_path)
            return
        except PermissionError:
            if attempt == attempts - 1:
                raise
            time.sleep(1)


def main():
    with open(ROSTER_PATH) as f:
        roster = json.load(f)
    with open(ATTENDANCE_PATH) as f:
        attendance = json.load(f)
    with open(SCHEDULES_PATH) as f:
        schedules = json.load(f)

    roster_rows = [r for r in roster.get('records', []) if r.get('employeeEmail')]
    roster_rows.sort(key=lambda r: (r.get('teamLeadName') or '', r.get('employeeName') or ''))
    roster_by_email = {clean_email(r['employeeEmail']): r for r in roster_rows}

    years = set()
    for period_key in (attendance.get('periods') or {}).keys():
        years.add(period_key.split('|')[0][:4])
    for email_entries in (attendance.get('autoEntries') or {}).values():
        for d in email_entries.keys():
            years.add(d[:4])
    if not years:
        years = {str(date_cls.today().year)}

    os.makedirs(OUT_ROOT, exist_ok=True)
    for year in sorted(years):
        months_with_data = set()
        for m in range(1, 13):
            dim = calendar.monthrange(int(year), m)[1]
            month_start, month_end = f'{year}-{m:02d}-01', f'{year}-{m:02d}-{dim:02d}'
            found = False
            for emp in roster_rows:
                for d in date_range(month_start, month_end):
                    if attendance_code_on_date(attendance, emp['employeeEmail'], d):
                        found = True
                        break
                if found:
                    break
            if found:
                months_with_data.add(m)

        wb = Workbook()
        wb.remove(wb.active)
        summary_ws = wb.create_sheet('Summary')
        build_summary_sheet(summary_ws, year, roster_rows, schedules, attendance, roster_by_email, months_with_data)
        for m in sorted(months_with_data):
            ws = wb.create_sheet(f'{MONTH_NAMES[m - 1]} {year}')
            build_month_sheet(ws, year, m, roster_rows, schedules, attendance, roster_by_email)
        if not months_with_data:
            today = date_cls.today()
            ws = wb.create_sheet(f'{MONTH_NAMES[today.month - 1]} {year}')
            build_month_sheet(ws, year, today.month, roster_rows, schedules, attendance, roster_by_email)
        wb.active = 0

        out_path = os.path.join(OUT_ROOT, f'{year}.xlsx')
        save_with_retry(wb, out_path)

    print(f'Exported consolidated attendance for {len(years)} year(s) under {OUT_ROOT}')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'consolidated attendance export failed: {e}', file=sys.stderr)
        sys.exit(1)
