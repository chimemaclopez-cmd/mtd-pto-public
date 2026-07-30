#!/usr/bin/env python3
# Regenerates the consolidated roster workbook whenever roster.json changes.
# Called automatically by zendesk-proxy.js after every roster save (create/update/
# delete/team-lead sync) - this is a one-way mirror (portal -> Excel). Edits made
# directly in the spreadsheet are never read back; the portal stays authoritative.
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROSTER_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'roster.json')
OUT_PATH = '/Users/mac/Library/CloudStorage/OneDrive-MoatableInc/Lofty PH Repository - Documents/Lofty TSR/Lofty TSR Roster.xlsx'

KPI_TIER = {'Senior TSR': 0, 'Voice Jr TSR': 1, 'Non-Voice Jr TSR': 1, 'Trainee': 2, 'Excluded': 3}

COLUMNS = [
    ('Employee Name', 'employeeName'),
    ('Employee ID', 'employeeId'),
    ('Email', 'employeeEmail'),
    ('Reports To', 'teamLeadName'),
    ('Reports To Email', 'teamLeadEmail'),
    ('Primary Channel', 'primaryChannel'),
    ('KPI Type', 'kpiType'),
    ('Employment Status', 'employmentStatus'),
    ('Active', 'active'),
    ('Hire Date', 'hireDate'),
    ('Birthday', 'birthday'),
    ('Effective Date', 'effectiveDate'),
    ('Separation Date', 'separationDate'),
    ('Senior TSR Assignment', 'seniorTsrAssignment'),
    ('Schedule Group', 'scheduleGroup'),
    ('Contact Number', 'contactNumber'),
    ('Contact Email', 'contactEmail'),
    ('Emergency Contact Name', 'emergencyContactName'),
    ('Emergency Contact Relationship', 'emergencyContactRelationship'),
    ('Emergency Contact Number', 'emergencyContactNumber'),
    ('Current Residence', 'currentResidence'),
    ('Notes', 'notes'),
    ('Last Updated', 'lastUpdated'),
]
COLUMN_WIDTHS = [22, 14, 30, 22, 26, 16, 18, 16, 8, 12, 12, 14, 14, 22, 16, 16, 26, 22, 22, 20, 30, 34, 20]


def is_leadership(r):
    return r.get('primaryChannel') == 'Leadership' or r.get('kpiType') == 'Excluded'


def sort_key(r):
    return (
        0 if is_leadership(r) else 1,
        '' if is_leadership(r) else (r.get('teamLeadName') or ''),
        KPI_TIER.get(r.get('kpiType'), 9),
        r.get('employeeName') or '',
    )


def main():
    with open(ROSTER_PATH) as f:
        data = json.load(f)
    records = sorted(data['records'], key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Roster'

    body_font = Font(name='Calibri', size=11)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2E46B8')
    leadership_fill = PatternFill('solid', fgColor='EAF0FF')
    thin = Side(style='thin', color='DFE2EA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    for row_idx, record in enumerate(records, start=2):
        leadership_row = is_leadership(record)
        for col_idx, (_, field) in enumerate(COLUMNS, start=1):
            value = record.get(field, '')
            if field == 'active':
                value = 'Yes' if value is not False else 'No'
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value not in (None, '') else '')
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if leadership_row:
                cell.fill = leadership_fill

    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{len(records) + 1}'

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f'Exported {len(records)} records to {OUT_PATH}')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'roster export failed: {e}', file=sys.stderr)
        sys.exit(1)
