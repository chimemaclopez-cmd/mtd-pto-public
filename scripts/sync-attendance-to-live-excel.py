#!/usr/bin/env python3
# One-way mirror: portal attendance.json -> the real "Lofty All Access File 2025.xlsx"
# company file (synced locally via the OneDrive desktop client). Called automatically
# by zendesk-proxy.js's saveAttendance() after every attendance write (manual entry,
# PTO approval/reversal, team-lead cloud-queue sync) - same trigger point and per-email
# scoping convention as export-attendance-xlsx.py.
#
# Deliberately conservative, given this is a shared, heavily-formatted, actively-used
# file that a human also opens directly in Excel:
#   - Only ever touches the sheet matching the CURRENT calendar month. It never creates
#     a month sheet - that's a manual, once-a-month step (see build_september_sheet.py
#     in the session scratchpad for the sheet-construction script). If that month's
#     sheet doesn't exist yet, this exits quietly and does nothing.
#   - Only writes day cells for employees already present in roster.json AND matched by
#     exact name to a row in that sheet. Every other row (the ~30 cross-functional /
#     leadership staff this portal doesn't track) is left completely untouched.
#   - Only ever writes a NON-EMPTY computed value. It never blanks a cell, since the
#     portal has no concept of "delete this attendance entry" - a missing/blank source
#     value is treated as "nothing to say about this date", not "clear whatever is
#     there."
#   - Skips silently (log line, exit 0, retried automatically on the next save) if the
#     file is currently open in Excel, detected via the `~$<name>` lock file Excel/Office
#     creates for any open file. It never fights an open editing session.
#   - Only calls wb.save() if at least one cell value actually changed, to keep the
#     lifetime number of full-workbook resaves (each of which fully rewrites every one
#     of the file's 36 sheets) as low as possible.
#   - Keeps one rolling pre-write backup (overwritten each time, not accumulated) so any
#     bad write can be recovered from immediately.
import json
import os
import shutil
import sys
import time
from datetime import date as date_cls, timedelta

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROSTER_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'roster.json')
ATTENDANCE_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'attendance.json')
BACKUP_PATH = os.path.join(SCRIPT_DIR, '..', 'data', '.backups', 'lofty-all-access-file-pre-sync.xlsx')

DEFAULT_LIVE_PATH = '/Users/mac/Library/CloudStorage/OneDrive-MoatableInc/Lofty All Access File 2025.xlsx'
LIVE_PATH = os.environ.get('LOFTY_ALL_ACCESS_FILE_PATH', DEFAULT_LIVE_PATH)

FULL_MONTH_NAMES = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                     'August', 'September', 'October', 'November', 'December']

FIRST_DATA_ROW = 5
LAST_DATA_ROW_SCAN = 90  # generous upper bound; real sheets end well before this
FIRST_DAY_COLUMN = 5     # column E

# roster.json's employeeName is the legal/full name used across the portal; the live
# Excel file's Name column was typed by hand over time and a few people are entered
# under the nickname/short form they actually go by. Known, confirmed mappings only -
# anyone else who doesn't match exactly is logged and skipped rather than guessed at.
NAME_ALIASES = {
    'Al John Kiven Sandoval': 'AJ Sandoval',
    'Nikcolai Oliver Acosta': 'Nikcolai Acosta',
}


def lock_file_path(live_path):
    directory, filename = os.path.split(live_path)
    return os.path.join(directory, f'~${filename}')


def resolve_code_for_date(attendance, email, date_str):
    # Mirrors server/pto-logic.js's attendanceCodeOnDate/attendanceMinutesLateOnDate/
    # attendanceLocationOnDate exactly: autoEntries (PTO-protected) wins unconditionally;
    # otherwise the most-specific/most-recently-saved matching period wins.
    auto = (attendance.get('autoEntries') or {}).get(email, {}).get(date_str)
    if auto:
        status = auto.get('status') if isinstance(auto, dict) else auto
        return status, auto if isinstance(auto, dict) else None
    month = date_str[:7]
    matches = sorted(
        [(k, v) for k, v in (attendance.get('periods') or {}).items()
         if k.startswith(month + '|') and k.split('|')[1] >= date_str],
        key=lambda kv: kv[0], reverse=True,
    )
    for _, period in matches:
        value = (period.get(email) or {}).get(date_str)
        if value:
            status = value.get('status') if isinstance(value, dict) else value
            return status, value if isinstance(value, dict) else None
    return None, None


def sheet_understands_late_detail(ws):
    # The "LATE (Location, Nm)" convention only means something to a sheet built with a
    # LATE-specific conditional-formatting rule (added starting with September 2026 - see
    # build_september_sheet.py). Older sheets (June/July/August 2026 and earlier) only have
    # RD/NCNS/PTO/SL/Onsite/WFH/EL rules with no LATE rule and no stopIfTrue precedence, so
    # a cell reading "LATE (Onsite, 80m)" would get colored as a plain Onsite day there
    # (it contains the substring "Onsite") - misleading. Detect and fall back to a bare
    # "LATE" on those sheets instead, so this script never needs its own list of which
    # months were upgraded.
    for cf_rules in ws.conditional_formatting._cf_rules.values():
        for rule in cf_rules:
            if rule.formula and any('LATE' in str(f) for f in rule.formula):
                return True
    return False


def format_cell_value(status, extra, late_detail_supported):
    if status != 'LATE':
        return status
    if not late_detail_supported:
        return 'LATE'
    minutes = extra.get('minutesLate') if isinstance(extra, dict) else None
    location = extra.get('location') if isinstance(extra, dict) else None
    loc_display = 'WFH' if location == 'WFH' else 'Onsite'
    if isinstance(minutes, (int, float)) and minutes > 0:
        return f'LATE ({loc_display}, {int(minutes)}m)'
    return 'LATE'


def month_days(year, month):
    d = date_cls(year, month, 1)
    out = []
    while d.month == month:
        out.append(d.isoformat())
        d += timedelta(days=1)
    return out


def find_employee_rows(ws):
    by_name = {}
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW_SCAN + 1):
        team_lead = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        if not team_lead or not name:
            continue
        name = str(name).strip()
        if name.lower() == 'name':
            continue  # the repeated mini-header row between the two roster sections
        by_name.setdefault(name, row)
    return by_name


def find_day_columns(ws, year, month):
    by_date = {}
    col = FIRST_DAY_COLUMN
    while True:
        value = ws.cell(row=3, column=col).value
        if value is None:
            if col > FIRST_DAY_COLUMN + 31:
                break
            col += 1
            continue
        if hasattr(value, 'year'):
            if value.year == year and value.month == month:
                by_date[value.date().isoformat() if hasattr(value, 'date') else value.isoformat()] = col
            elif by_date:
                break  # ran past the end of this month's day-grid
        col += 1
        if col > FIRST_DAY_COLUMN + 40:
            break
    return by_date


def save_with_retry(wb, out_path, attempts=3):
    last_error = None
    for attempt in range(attempts):
        try:
            wb.save(out_path)
            return
        except (TimeoutError, OSError) as e:
            last_error = e
            if attempt < attempts - 1:
                time.sleep(1.5 * (attempt + 1))
    raise last_error


def main():
    target_emails = {e.strip().lower() for e in sys.argv[1:] if e.strip()}

    if os.path.exists(lock_file_path(LIVE_PATH)):
        print('[excel-sync] Live file is currently open in Excel - skipping this cycle, will retry on next attendance save.')
        return 0
    if not os.path.exists(LIVE_PATH):
        print(f'[excel-sync] Live file not found at {LIVE_PATH} - skipping.', file=sys.stderr)
        return 1

    roster = json.load(open(ROSTER_PATH))
    attendance = json.load(open(ATTENDANCE_PATH))

    employees = [r for r in roster.get('records', []) if r.get('active')]
    if target_emails:
        employees = [r for r in employees if str(r.get('employeeEmail', '')).strip().lower() in target_emails]
    if not employees:
        print('[excel-sync] No matching active employees to sync - skipping.')
        return 0

    today = date_cls.today()
    sheet_name = f'{FULL_MONTH_NAMES[today.month - 1]} {today.year}'

    wb = load_workbook(LIVE_PATH)
    if sheet_name not in wb.sheetnames:
        print(f'[excel-sync] Sheet "{sheet_name}" does not exist yet in the live file - skipping (create it first).')
        return 0
    ws = wb[sheet_name]

    name_to_row = find_employee_rows(ws)
    date_to_col = find_day_columns(ws, today.year, today.month)
    if not date_to_col:
        print(f'[excel-sync] Could not locate any day columns for {sheet_name} - skipping.', file=sys.stderr)
        return 1
    late_detail_supported = sheet_understands_late_detail(ws)

    dates_this_month = [d for d in month_days(today.year, today.month) if d <= today.isoformat()]

    changed = 0
    unmatched_names = set()
    for emp in employees:
        email = str(emp.get('employeeEmail', '')).strip().lower()
        name = str(emp.get('employeeName', '')).strip()
        row = name_to_row.get(name) or name_to_row.get(NAME_ALIASES.get(name, ''))
        if not row:
            unmatched_names.add(name)
            continue
        for date_str in dates_this_month:
            col = date_to_col.get(date_str)
            if not col:
                continue
            status, extra = resolve_code_for_date(attendance, email, date_str)
            if not status:
                continue
            new_value = format_cell_value(status, extra, late_detail_supported)
            cell = ws.cell(row=row, column=col)
            if cell.value != new_value:
                cell.value = new_value
                changed += 1

    if unmatched_names:
        print(f'[excel-sync] {len(unmatched_names)} roster name(s) not found as a row in "{sheet_name}": {", ".join(sorted(unmatched_names))}', file=sys.stderr)

    if changed == 0:
        print(f'[excel-sync] {sheet_name}: no new attendance values to write.')
        return 0

    os.makedirs(os.path.dirname(BACKUP_PATH), exist_ok=True)
    shutil.copyfile(LIVE_PATH, BACKUP_PATH)

    save_with_retry(wb, LIVE_PATH)
    print(f'[excel-sync] {sheet_name}: wrote {changed} cell(s).')
    return 0


if __name__ == '__main__':
    sys.exit(main())
